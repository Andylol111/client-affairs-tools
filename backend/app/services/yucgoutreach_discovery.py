"""
YUCGoutreach company discovery: research per-company email patterns, employee seeds
(LinkedIn via Apify or web search), parallel enrichment (default 4 workers), AI scoring.
Stores rows in yucgoutreach_prospects for SQL queries and Excel export.
"""
from __future__ import annotations

import asyncio
import json
import os
import re
from typing import Any

import httpx

from app.database import get_db, row_to_dict
from app.services.contact_scraper import is_employee_outreach_email, normalize_domain
from app.services.linkedin_scraper import scrape_linkedin_company

OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.2")


async def _ollama_json(prompt: str, timeout: float = 90.0) -> dict[str, Any] | None:
    """Ask Ollama for a single JSON object (async HTTP)."""
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            r = await client.post(
                f"{OLLAMA_URL.rstrip('/')}/api/chat",
                json={
                    "model": OLLAMA_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "stream": False,
                    "options": {"temperature": 0.3},
                },
            )
            if r.status_code != 200:
                return None
            body = r.json()
            content = (body.get("message") or {}).get("content") or body.get("response") or ""
            content = str(content).strip()
            m = re.search(r"\{[\s\S]*\}", content)
            if not m:
                return None
            return json.loads(m.group())
    except Exception:
        return None


async def _tavily_search(query: str, max_results: int = 8) -> list[dict[str, Any]]:
    key = (os.getenv("TAVILY_API_KEY") or "").strip()
    if not key:
        return []
    try:
        async with httpx.AsyncClient(timeout=35.0) as client:
            r = await client.post(
                "https://api.tavily.com/search",
                json={
                    "api_key": key,
                    "query": query,
                    "search_depth": "basic",
                    "max_results": max_results,
                },
            )
            r.raise_for_status()
            data = r.json()
    except Exception:
        return []
    out = []
    for x in data.get("results") or []:
        out.append(
            {
                "title": x.get("title") or "",
                "url": x.get("url") or "",
                "content": (x.get("content") or "")[:1200],
            }
        )
    return out


def _split_first_last(full_name: str) -> tuple[str, str]:
    name = (full_name or "").strip()
    if not name:
        return "", ""
    parts = [p for p in re.split(r"[\s,]+", name) if p]
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[-1]


def _apply_local_template(tpl: str, first: str, last: str) -> str:
    f = (first or "").lower().strip()
    l = (last or "").lower().strip()
    fi = f[:1] if f else ""
    li = l[:1] if l else ""
    s = tpl
    s = s.replace("{first}", f).replace("{last}", l).replace("{f}", fi).replace("{l}", li)
    s = s.replace("{FIRST}", f.upper()).replace("{LAST}", l.upper())
    return re.sub(r"[^a-z0-9._+-]", "", s, flags=re.I)


def _predict_emails(
    first: str,
    last: str,
    domain: str,
    templates: list[str],
) -> list[str]:
    domain = normalize_domain(domain)
    if not domain:
        return []
    seen: set[str] = set()
    emails: list[str] = []
    for tpl in templates:
        local = _apply_local_template(tpl, first, last)
        if not local or "." not in local and len(local) < 2:
            continue
        email = f"{local}@{domain}".lower()
        if email in seen:
            continue
        seen.add(email)
        if is_employee_outreach_email(email):
            emails.append(email)
    return emails


async def _research_email_patterns(
    company_name: str,
    domain: str,
    research_snippets: list[dict[str, Any]],
) -> dict[str, Any]:
    snippet_text = "\n\n".join(
        f"[{i+1}] {s.get('title', '')} {s.get('url', '')}\n{s.get('content', '')}"
        for i, s in enumerate(research_snippets[:10])
    )
    dom = normalize_domain(domain) or "unknown"
    prompt = f"""You are a B2B data researcher. Based on the company "{company_name}" and domain hint "{dom}",
infer how employee work emails are MOST LIKELY formatted at THIS company specifically.
Use the web snippets when helpful; if snippets are thin, infer from common US corporate patterns but lower confidence.

Snippets:
{snippet_text if snippet_text.strip() else "(no snippets — infer conservative templates)"}

Return ONLY valid JSON:
{{
  "local_part_templates": ["{{first}}.{{last}}", "{{f}}{{last}}"],
  "reasoning": "one short sentence",
  "pattern_confidence": 0.0
}}
Rules for local_part_templates: use placeholders exactly {{first}}, {{last}}, {{f}}, {{l}} (lowercase).
Provide 3–6 templates ordered by likelihood for THIS company. pattern_confidence is 0–1."""

    data = await _ollama_json(prompt, timeout=120.0)
    out: dict[str, Any] = {
        "local_part_templates": ["{first}.{last}", "{f}{last}", "{first}{last}"],
        "reasoning": "Default patterns (research unavailable)",
        "pattern_confidence": 0.35,
    }
    if data and isinstance(data.get("local_part_templates"), list):
        tpls = [str(t) for t in data["local_part_templates"] if t]
        if tpls:
            out["local_part_templates"] = tpls
        if data.get("reasoning"):
            out["reasoning"] = str(data["reasoning"])[:500]
        try:
            out["pattern_confidence"] = float(data.get("pattern_confidence", 0.5))
        except (TypeError, ValueError):
            pass
    return out


async def _infer_domain(company_name: str) -> str:
    """Best-effort domain from web when user did not provide one."""
    results = await _tavily_search(f"{company_name} official company website homepage", max_results=5)
    if not results:
        return ""
    snippet = "\n".join(f"{r.get('title')} {r.get('url')} {r.get('content', '')[:200]}" for r in results)
    data = await _ollama_json(
        f"""From the snippets, what is the primary corporate website domain for "{company_name}"?
Return ONLY JSON: {{"domain": "example.com"}} Use bare domain, no protocol. If unknown: {{"domain": ""}}.

Snippets:
{snippet[:4000]}""",
        timeout=45.0,
    )
    if not data:
        url = (results[0].get("url") or "") if results else ""
        return normalize_domain(url)
    return normalize_domain(str(data.get("domain") or ""))


async def _company_meta(company_name: str, domain: str) -> dict[str, str]:
    results = await _tavily_search(
        f"{company_name} company industry headquarters employee count {domain}".strip(),
        max_results=6,
    )
    if not results:
        return {"country": "", "employees": "", "industry": "", "keywords": "", "keywords_2": ""}
    snippet_text = "\n".join(f"{r.get('title')} — {r.get('content', '')[:400]}" for r in results[:5])
    prompt = f"""From the text about "{company_name}", extract fields. Return ONLY JSON:
{{"country": "", "employees": "", "industry": "", "keywords_1": "", "keywords_2": ""}}
Use short strings; employees like "500-1000" or "unknown" if unclear."""
    data = await _ollama_json(
        prompt + "\n\nSource text:\n" + snippet_text[:3500],
        timeout=60.0,
    )
    if not data:
        return {"country": "", "employees": "", "industry": "", "keywords": "", "keywords_2": ""}
    return {
        "country": str(data.get("country") or ""),
        "employees": str(data.get("employees") or ""),
        "industry": str(data.get("industry") or ""),
        "keywords": str(data.get("keywords_1") or data.get("keywords") or ""),
        "keywords_2": str(data.get("keywords_2") or ""),
    }


async def _tavily_employee_seeds(company_name: str, max_n: int) -> list[dict[str, Any]]:
    results = await _tavily_search(
        f'{company_name} employees OR leadership OR "works at" site:linkedin.com/in',
        max_results=12,
    )
    if not results:
        return []
    snippet_text = "\n\n".join(
        f"[{i+1}] {r.get('title', '')}\n{r.get('url', '')}\n{r.get('content', '')}"
        for i, r in enumerate(results)
    )
    prompt = f"""Extract up to {max_n} DISTINCT people who appear to work at or be associated with "{company_name}".
Prefer clear full names. Return ONLY JSON:
{{"people": [{{"full_name": "", "title": "", "linkedin_url": ""}}]}}
linkedin_url must be linkedin.com/in/... when present, else ""."""
    data = await _ollama_json(prompt + "\n\nSearch results:\n" + snippet_text[:6000], timeout=120.0)
    people = (data or {}).get("people") if isinstance(data, dict) else None
    if not isinstance(people, list):
        return []
    seeds = []
    for p in people:
        if not isinstance(p, dict):
            continue
        fn = str(p.get("full_name") or p.get("name") or "").strip()
        if len(fn) < 3:
            continue
        seeds.append(
            {
                "full_name": fn,
                "title": str(p.get("title") or p.get("headline") or "").strip(),
                "linkedin_url": str(p.get("linkedin_url") or "").strip(),
            }
        )
        if len(seeds) >= max_n:
            break
    return seeds


def _parse_secondary_score(data: dict[str, Any] | None, fallback: float) -> float:
    if not data:
        return fallback
    v = data.get("yucgoutreach_score")
    if v is None:
        v = data.get("apollo_score")
    try:
        return float(v)
    except (TypeError, ValueError):
        return fallback


async def _score_prospect_llm(
    company_name: str,
    full_name: str,
    title: str,
    email_guess: str,
    linkedin_url: str,
    pattern_confidence: float,
    resume_snippets: str,
) -> dict[str, Any]:
    prompt = f"""Score this B2B prospect for outreach data quality.

Company: {company_name}
Person: {full_name}
Title: {title}
Predicted email: {email_guess}
LinkedIn: {linkedin_url}
Email pattern confidence (0-1): {pattern_confidence}
Web evidence (may be partial):
{resume_snippets[:2000]}

Return ONLY JSON:
{{
  "score": 0,
  "yucgoutreach_score": 0,
  "fit_status": "unknown",
  "qualification_notes": "",
  "verified_hint": 0
}}
score and yucgoutreach_score are 0-100 integers (yucgoutreach_score is a secondary lead-quality signal; both are in-house, not from any vendor API).
fit_status one of: strong, medium, weak, unknown.
verified_hint: 1 if email guess is plausibly corroborated by evidence, else 0."""
    data = await _ollama_json(prompt, timeout=90.0)
    if not data:
        base = 40.0 + 30.0 * pattern_confidence
        if linkedin_url:
            base += 15
        return {
            "score": min(100.0, base),
            "yucgoutreach_score": min(100.0, base - 5),
            "fit_status": "medium" if base >= 55 else "weak",
            "qualification_notes": "Heuristic score (LLM unavailable).",
            "verified_hint": 1 if linkedin_url else 0,
        }
    try:
        score = float(data.get("score", 50))
    except (TypeError, ValueError):
        score = 50.0
    secondary = _parse_secondary_score(data, score)
    return {
        "score": max(0.0, min(100.0, score)),
        "yucgoutreach_score": max(0.0, min(100.0, secondary)),
        "fit_status": str(data.get("fit_status") or "unknown"),
        "qualification_notes": str(data.get("qualification_notes") or "")[:2000],
        "verified_hint": 1
        if data.get("verified_hint") in (1, True, "1", "true", "True")
        else 0,
    }


async def _run_update(
    run_id: int,
    *,
    status: str | None = None,
    progress_pct: float | None = None,
    progress_message: str | None = None,
    prospects_count: int | None = None,
    research_json: str | None = None,
    error_message: str | None = None,
    completed: bool = False,
):
    db = await get_db()
    try:
        sets = ["updated_at = CURRENT_TIMESTAMP"]
        args: list[Any] = []
        if status is not None:
            sets.append("status = ?")
            args.append(status)
        if progress_pct is not None:
            sets.append("progress_pct = ?")
            args.append(progress_pct)
        if progress_message is not None:
            sets.append("progress_message = ?")
            args.append(progress_message[:2000])
        if prospects_count is not None:
            sets.append("prospects_count = ?")
            args.append(prospects_count)
        if research_json is not None:
            sets.append("research_json = ?")
            args.append(research_json)
        if error_message is not None:
            sets.append("error_message = ?")
            args.append(error_message[:2000])
        if completed:
            sets.append("completed_at = CURRENT_TIMESTAMP")
        args.append(run_id)
        await db.execute(
            f"UPDATE yucgoutreach_discovery_runs SET {', '.join(sets)} WHERE id = ?",
            tuple(args),
        )
        await db.commit()
    finally:
        await db.close()


async def execute_yucgoutreach_run(run_id: int) -> None:
    db = await get_db()
    try:
        cur = await db.execute(
            "SELECT * FROM yucgoutreach_discovery_runs WHERE id = ?",
            (run_id,),
        )
        row = await cur.fetchone()
        if not row:
            return
        spec = row_to_dict(row)
    finally:
        await db.close()

    company = (spec.get("company_name") or "").strip()
    domain_in = (spec.get("company_domain") or "").strip()
    linkedin_url = (spec.get("linkedin_company_url") or "").strip() or None
    max_prospects = int(spec.get("max_prospects") or 25)
    max_prospects = max(1, min(max_prospects, 200))
    workers = int(spec.get("worker_concurrency") or 4)
    workers = max(1, min(workers, 16))
    domain = normalize_domain(domain_in) if domain_in else ""

    await _run_update(
        run_id,
        status="running",
        progress_pct=2.0,
        progress_message="Researching company email conventions (per-company patterns)…",
    )

    research_snippets: list[dict[str, Any]] = []
    if domain:
        research_snippets.extend(
            await _tavily_search(
                f'"{company}" employee email format @{domain} OR "e-mail" site:{domain}',
                max_results=8,
            )
        )
    research_snippets.extend(
        await _tavily_search(f"{company} press contact OR media contact email", max_results=4)
    )

    pattern_pack = await _research_email_patterns(company, domain, research_snippets)
    templates_raw = pattern_pack.get("local_part_templates") or []
    templates = []
    for raw in templates_raw:
        t = str(raw).strip()
        t = re.sub(r"\{\{", "{", t)
        t = re.sub(r"\}\}", "}", t)
        if "{first}" in t or "{last}" in t or "{f}" in t or "{l}" in t:
            templates.append(t)

    if not templates:
        templates = ["{first}.{last}", "{f}{last}", "{first}{last}"]

    if not domain:
        domain = await _infer_domain(company)
        if domain:
            await _run_update(
                run_id,
                progress_message=f"Inferred domain {domain} — refining email pattern…",
            )
            research_snippets.extend(
                await _tavily_search(
                    f'"{company}" employee email @{domain}',
                    max_results=5,
                )
            )
            pattern_pack = await _research_email_patterns(company, domain, research_snippets)
            tr2 = pattern_pack.get("local_part_templates") or []
            templates = []
            for raw in tr2:
                t = str(raw).strip()
                t = re.sub(r"\{\{", "{", t)
                t = re.sub(r"\}\}", "}", t)
                if "{first}" in t or "{last}" in t or "{f}" in t or "{l}" in t:
                    templates.append(t)
            if not templates:
                templates = ["{first}.{last}", "{f}{last}", "{first}{last}"]

    pattern_confidence = float(pattern_pack.get("pattern_confidence") or 0.5)

    await _run_update(
        run_id,
        progress_pct=18.0,
        progress_message="Loading company profile (industry, geography)…",
        research_json=json.dumps(
            {"email_pattern": pattern_pack, "snippet_count": len(research_snippets)}
        ),
    )

    meta = await _company_meta(company, domain)
    kw1 = meta.get("keywords") or ""
    kw2 = meta.get("keywords_2") or ""

    await _run_update(run_id, progress_pct=28.0, progress_message="Discovering named employees…")

    seeds: list[dict[str, Any]] = []
    li_error: str | None = None
    if linkedin_url:
        li = await scrape_linkedin_company(linkedin_url, max_employees=max_prospects)
        li_error = li.get("error")
        for c in li.get("contacts") or []:
            nm = (c.get("name") or "").strip()
            if not nm:
                continue
            seeds.append(
                {
                    "full_name": nm,
                    "title": (c.get("title") or "")[:300],
                    "linkedin_url": (c.get("linkedin_url") or "").strip(),
                }
            )
        if not seeds and li_error:
            await _run_update(
                run_id,
                progress_message=f"LinkedIn step: {li_error[:200]}. Trying web search for names…",
            )

    if len(seeds) < max_prospects:
        need = max_prospects - len(seeds)
        seeds.extend(await _tavily_employee_seeds(company, need))

    # De-dupe by name
    seen_names: set[str] = set()
    unique_seeds: list[dict[str, Any]] = []
    for s in seeds:
        key = s["full_name"].lower()
        if key in seen_names:
            continue
        seen_names.add(key)
        unique_seeds.append(s)
        if len(unique_seeds) >= max_prospects:
            break
    seeds = unique_seeds

    if not seeds:
        await _run_update(
            run_id,
            status="completed",
            progress_pct=100.0,
            progress_message="No named employees found. Add a LinkedIn company URL and APIFY_API_TOKEN, or set TAVILY_API_KEY for web-based name discovery.",
            prospects_count=0,
            completed=True,
        )
        return

    sem = asyncio.Semaphore(workers)
    inserted = 0
    n = len(seeds)
    lock = asyncio.Lock()

    async def enrich_one(idx: int, seed: dict[str, Any]) -> None:
        nonlocal inserted
        full_name = seed["full_name"]
        first, last = _split_first_last(full_name)
        if not last:
            first, last = full_name, ""
        emails = _predict_emails(first, last, domain, templates)
        email_pick = emails[0] if emails else ""

        person_q = f"{full_name} {company} email OR resume OR contact"
        evidence = await _tavily_search(person_q, max_results=5)
        ev_text = "\n".join(f"{e.get('title')} {e.get('content', '')[:350]}" for e in evidence)

        async with sem:
            scores = await _score_prospect_llm(
                company,
                full_name,
                seed.get("title") or "",
                email_pick,
                seed.get("linkedin_url") or "",
                pattern_confidence,
                ev_text,
            )

            linkedin = seed.get("linkedin_url") or ""
            contact_profile = linkedin or (evidence[0].get("url") if evidence else "") or ""
            verified = 0
            if scores.get("verified_hint"):
                verified = 1
            elif email_pick and email_pick.lower() in ev_text.lower():
                verified = 1
            elif linkedin and "linkedin.com/in/" in linkedin:
                verified = 1

            evidence_obj = {
                "pattern_templates": templates,
                "predicted_emails_tried": emails[:8],
                "search_hits": [{"title": e.get("title"), "url": e.get("url")} for e in evidence[:5]],
            }
            dbi = await get_db()
            try:
                await dbi.execute(
                    """INSERT INTO yucgoutreach_prospects (
                        run_id, first_name, last_name, email, company, contact_url, title,
                        account_url, photo_url, account_link, phone, phone_code, verified,
                        qualification_notes, contact_profile_url, linkedin_url, fit_status,
                        score, country, employees, industry, keywords_1, keywords_2,
                        yucgoutreach_score, evidence_json
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        run_id,
                        first,
                        last,
                        email_pick or None,
                        company,
                        contact_profile or None,
                        (seed.get("title") or None),
                        None,
                        None,
                        None,
                        None,
                        None,
                        verified,
                        scores.get("qualification_notes"),
                        contact_profile or None,
                        linkedin or None,
                        scores.get("fit_status"),
                        scores.get("score"),
                        meta.get("country") or None,
                        meta.get("employees") or None,
                        meta.get("industry") or None,
                        kw1 or None,
                        kw2 or None,
                        scores.get("yucgoutreach_score"),
                        json.dumps(evidence_obj)[:8000],
                    ),
                )
                await dbi.commit()
            finally:
                await dbi.close()

        async with lock:
            inserted += 1
            ic = inserted

        pct = 30.0 + (idx + 1) / max(n, 1) * 65.0
        await _run_update(
            run_id,
            progress_pct=min(99.0, pct),
            progress_message=f"Enriched {idx + 1}/{n}: {full_name[:40]}",
            prospects_count=ic,
        )

    await asyncio.gather(*[enrich_one(i, s) for i, s in enumerate(seeds)])

    await _run_update(
        run_id,
        status="completed",
        progress_pct=100.0,
        progress_message=f"Done — {inserted} prospects saved (parallel workers: {workers}).",
        prospects_count=inserted,
        completed=True,
    )


async def _yucgoutreach_run_guard(run_id: int) -> None:
    try:
        await execute_yucgoutreach_run(run_id)
    except Exception as e:
        await _run_update(
            run_id,
            status="failed",
            progress_pct=100.0,
            error_message=str(e),
            completed=True,
        )


async def build_yucgoutreach_excel_bytes(run_id: int) -> bytes:
    """Build .xlsx for a run (async DB + sync openpyxl)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Company",
        "Contact URL",
        "Title",
        "Account URL",
        "Photo URL",
        "Account Link",
        "Phone",
        "Phone Code",
        "Verified",
        "Qualification Notes",
        "Contact Profile URL",
        "LinkedIn URL",
        "Fit Status",
        "Score",
        "Country",
        "Employees",
        "Industry",
        "Keywords 1",
        "Keywords 2",
        "YUCGoutreach Score",
    ]

    db = await get_db()
    try:
        cur = await db.execute(
            "SELECT company_name FROM yucgoutreach_discovery_runs WHERE id = ?",
            (run_id,),
        )
        r0 = await cur.fetchone()
        if not r0:
            raise ValueError("Run not found")
        company_name = r0["company_name"]
        cur = await db.execute(
            """SELECT first_name, last_name, email, company, contact_url, title,
            account_url, photo_url, account_link, phone, phone_code, verified,
            qualification_notes, contact_profile_url, linkedin_url, fit_status,
            score, country, employees, industry, keywords_1, keywords_2, yucgoutreach_score
            FROM yucgoutreach_prospects WHERE run_id = ? ORDER BY id""",
            (run_id,),
        )
        rows = await cur.fetchall()
        prospect_rows = [row_to_dict(r) for r in rows]
    finally:
        await db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font

    for ri, pr in enumerate(prospect_rows, 2):
        secondary = pr.get("yucgoutreach_score")
        if secondary is None:
            secondary = pr.get("apollo_score")
        ws.cell(row=ri, column=1, value=pr.get("first_name"))
        ws.cell(row=ri, column=2, value=pr.get("last_name"))
        ws.cell(row=ri, column=3, value=pr.get("email"))
        ws.cell(row=ri, column=4, value=pr.get("company") or company_name)
        ws.cell(row=ri, column=5, value=pr.get("contact_url"))
        ws.cell(row=ri, column=6, value=pr.get("title"))
        ws.cell(row=ri, column=7, value=pr.get("account_url"))
        ws.cell(row=ri, column=8, value=pr.get("photo_url"))
        ws.cell(row=ri, column=9, value=pr.get("account_link"))
        ws.cell(row=ri, column=10, value=pr.get("phone"))
        ws.cell(row=ri, column=11, value=pr.get("phone_code"))
        ws.cell(row=ri, column=12, value="Yes" if pr.get("verified") else "No")
        ws.cell(row=ri, column=13, value=pr.get("qualification_notes"))
        ws.cell(row=ri, column=14, value=pr.get("contact_profile_url"))
        ws.cell(row=ri, column=15, value=pr.get("linkedin_url"))
        ws.cell(row=ri, column=16, value=pr.get("fit_status"))
        ws.cell(row=ri, column=17, value=pr.get("score"))
        ws.cell(row=ri, column=18, value=pr.get("country"))
        ws.cell(row=ri, column=19, value=pr.get("employees"))
        ws.cell(row=ri, column=20, value=pr.get("industry"))
        ws.cell(row=ri, column=21, value=pr.get("keywords_1"))
        ws.cell(row=ri, column=22, value=pr.get("keywords_2"))
        ws.cell(row=ri, column=23, value=secondary)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
